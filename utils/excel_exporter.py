"""
Excel Test Case Exporter
Generates professional Excel test case documents
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict
from datetime import datetime
from agents.state import AgentState, TestCase


class ExcelExporter:
    """
    Exports test cases to a professional Excel format
    """
    
    def __init__(self):
        self.priority_colors = {
            "P0": "FF0000",  # Red
            "P1": "FFA500",  # Orange
            "P2": "FFFF00",  # Yellow
            "P3": "90EE90"   # Light Green
        }
    
    def export_test_cases(self, state: AgentState, output_path: str) -> str:
        """
        Export test cases to Excel
        
        Args:
            state: Final agent state with test cases
            output_path: Path to save Excel file
        
        Returns:
            Path to generated Excel file
        """
        wb = Workbook()
        
        # Create sheets
        self._create_summary_sheet(wb, state)
        self._create_test_cases_sheet(wb, state)
        self._create_qa_roadmap_sheet(wb, state)
        self._create_coverage_sheet(wb, state)
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Save
        wb.save(output_path)
        return output_path
    
    def _create_summary_sheet(self, wb: Workbook, state: AgentState):
        """Create summary overview sheet"""
        ws = wb.active
        ws.title = "Summary"
        
        # Header
        ws["A1"] = "Ticket-to-Test AI - QA Execution Summary"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells("A1:D1")
        
        # Ticket Info
        ticket = state["ticket_info"]
        row = 3
        info_fields = [
            ("Ticket ID:", ticket.get("ticket_id", "N/A")),
            ("Title:", ticket.get("title", "N/A")),
            ("Type:", ticket.get("ticket_type", "N/A")),
            ("Priority:", ticket.get("priority", "N/A")),
            ("Status:", ticket.get("status", "N/A")),
            ("", ""),
            ("Generated On:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Processing Time:", f"{state.get('processing_time', 0):.2f} seconds"),
        ]
        
        for label, value in info_fields:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1
        
        # Test Case Statistics
        row += 1
        ws[f"A{row}"] = "Test Case Statistics"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        row += 1
        
        test_cases = state["test_cases"]
        priority_counts = {}
        category_counts = {}
        
        for tc in test_cases:
            priority = tc.get("priority", "P2")
            category = tc.get("category", "Other")
            priority_counts[priority] = priority_counts.get(priority, 0) + 1
            category_counts[category] = category_counts.get(category, 0) + 1
        
        ws[f"A{row}"] = "Total Test Cases:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = len(test_cases)
        row += 1
        
        ws[f"A{row}"] = "By Priority:"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1
        
        for priority in ["P0", "P1", "P2", "P3"]:
            count = priority_counts.get(priority, 0)
            ws[f"B{row}"] = f"{priority}:"
            ws[f"C{row}"] = count
            ws[f"C{row}"].fill = PatternFill(
                start_color=self.priority_colors.get(priority, "FFFFFF"),
                end_color=self.priority_colors.get(priority, "FFFFFF"),
                fill_type="solid"
            )
            row += 1
        
        # Format column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
    
    def _create_test_cases_sheet(self, wb: Workbook, state: AgentState):
        """Create detailed test cases sheet"""
        ws = wb.create_sheet("Test Cases")
        
        # Headers
        headers = [
            "Test ID", "Priority", "Category", "Title", 
            "Preconditions", "Test Steps", "Expected Result",
            "Test Data", "Automation Feasibility"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Data
        test_cases = state["test_cases"]
        for row_idx, tc in enumerate(test_cases, 2):
            ws.cell(row_idx, 1, tc.get("test_id", ""))
            
            # Priority with color
            priority_cell = ws.cell(row_idx, 2, tc.get("priority", "P2"))
            priority = tc.get("priority", "P2")
            priority_cell.fill = PatternFill(
                start_color=self.priority_colors.get(priority, "FFFFFF"),
                end_color=self.priority_colors.get(priority, "FFFFFF"),
                fill_type="solid"
            )
            priority_cell.alignment = Alignment(horizontal="center")
            
            ws.cell(row_idx, 3, tc.get("category", ""))
            ws.cell(row_idx, 4, tc.get("title", ""))
            ws.cell(row_idx, 5, tc.get("preconditions", ""))
            
            # Test steps as numbered list
            steps = tc.get("test_steps", [])
            steps_text = "\n".join(f"{i+1}. {step}" for i, step in enumerate(steps))
            ws.cell(row_idx, 6, steps_text)
            
            ws.cell(row_idx, 7, tc.get("expected_result", ""))
            ws.cell(row_idx, 8, tc.get("test_data", ""))
            ws.cell(row_idx, 9, tc.get("automation_feasibility", ""))
            
            # Wrap text for all cells in this row
            for col in range(1, 10):
                ws.cell(row_idx, col).alignment = Alignment(wrap_text=True, vertical="top")
        
        # Format column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 50
        ws.column_dimensions["G"].width = 30
        ws.column_dimensions["H"].width = 20
        ws.column_dimensions["I"].width = 20
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    def _create_qa_roadmap_sheet(self, wb: Workbook, state: AgentState):
        """Create QA roadmap sheet"""
        ws = wb.create_sheet("QA Roadmap")
        
        # Header
        ws["A1"] = "QA Execution Roadmap"
        ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells("A1:C1")
        
        # Roadmap
        row = 3
        qa_roadmap = state.get("qa_roadmap", {})
        
        for category, scenarios in qa_roadmap.items():
            # Category header
            ws[f"A{row}"] = category
            ws[f"A{row}"].font = Font(size=12, bold=True)
            ws[f"A{row}"].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws.merge_cells(f"A{row}:C{row}")
            row += 1
            
            # Scenarios
            for scenario in scenarios:
                ws[f"A{row}"] = "•"
                ws[f"B{row}"] = scenario
                ws[f"B{row}"].alignment = Alignment(wrap_text=True)
                row += 1
            
            row += 1  # Empty row between categories
        
        # Column widths
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 80
        ws.column_dimensions["C"].width = 20
    
    def _create_coverage_sheet(self, wb: Workbook, state: AgentState):
        """Create coverage analysis sheet"""
        ws = wb.create_sheet("Coverage Analysis")
        
        # Header
        ws["A1"] = "Coverage Analysis & Gaps"
        ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells("A1:B1")
        
        row = 3
        
        # Requirements Coverage
        ws[f"A{row}"] = "Requirements to Cover"
        ws[f"A{row}"].font = Font(size=12, bold=True)
        row += 1
        
        for req in state.get("extracted_requirements", []):
            ws[f"A{row}"] = "•"
            ws[f"B{row}"] = req
            ws[f"B{row}"].alignment = Alignment(wrap_text=True)
            row += 1
        
        row += 1
        
        # Coverage Gaps
        ws[f"A{row}"] = "Identified Coverage Gaps"
        ws[f"A{row}"].font = Font(size=12, bold=True)
        ws[f"A{row}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        row += 1
        
        gaps = state.get("coverage_gaps", [])
        if gaps:
            for gap in gaps:
                ws[f"A{row}"] = "⚠"
                ws[f"B{row}"] = gap
                ws[f"B{row}"].alignment = Alignment(wrap_text=True)
                ws[f"B{row}"].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                row += 1
        else:
            ws[f"B{row}"] = "No coverage gaps identified - Excellent coverage!"
            ws[f"B{row}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            row += 1
        
        row += 1
        
        # Clarification Questions
        ws[f"A{row}"] = "Clarification Questions"
        ws[f"A{row}"].font = Font(size=12, bold=True)
        row += 1
        
        questions = state.get("clarification_questions", [])
        if questions:
            for question in questions:
                ws[f"A{row}"] = "?"
                ws[f"B{row}"] = question
                ws[f"B{row}"].alignment = Alignment(wrap_text=True)
                row += 1
        else:
            ws[f"B{row}"] = "No clarification needed"
            row += 1
        
        # Column widths
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 100
